import time
from contextlib import contextmanager
from unittest.mock import MagicMock

import pyarrow as pa
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from airflow_lite.api.app import create_app
from airflow_lite.config.settings import ApiConfig, PipelineConfig, WebUIConfig
from airflow_lite.export import FilesystemAnalyticsExportService
from airflow_lite.export.writers import XlsxExportWriter, XlsxRowLimitExceededError
from airflow_lite.query import DuckDBAnalyticsQueryService
from airflow_lite.query.contracts import (
    ExportCreateRequest,
    ExportFormat,
    ExportJobStatus,
)
from airflow_lite.query.service import AnalyticsExportPlan


class _BatchReader:
    def __init__(self, batches: list[pa.RecordBatch]):
        if not batches:
            raise ValueError("batches must not be empty")
        self._batches = list(batches)
        self.schema = self._batches[0].schema

    def __iter__(self):
        return iter(self._batches)


class _StaticPlanQueryProvider:
    def __init__(self, batches: list[pa.RecordBatch]):
        self._batches = list(batches)

    def build_export_plan(self, request: ExportCreateRequest) -> AnalyticsExportPlan:
        return AnalyticsExportPlan(
            dataset=request.dataset,
            action_key=request.action_key,
            format=request.format,
            sql="SELECT 1",
            params=[],
            file_stem="test-export",
        )

    @contextmanager
    def execute_export_batches(self, sql: str, params: list, rows_per_batch: int):
        yield _BatchReader(self._batches)


def _build_test_batch(row_count: int) -> pa.RecordBatch:
    return pa.record_batch(
        [
            pa.array([f"SRC_{index}" for index in range(row_count)], type=pa.string()),
            pa.array(list(range(row_count)), type=pa.int64()),
        ],
        names=["source_name", "row_count"],
    )


def _wait_for_service_job(
    export_service: FilesystemAnalyticsExportService, job_id: str
):
    for _ in range(60):
        response = export_service.get_job(job_id)
        if response.status in {ExportJobStatus.COMPLETED, ExportJobStatus.FAILED}:
            return response
        time.sleep(0.05)
    raise AssertionError(f"export job did not finish in time: {job_id}")


def _wait_for_api_job(client: TestClient, job_id: str) -> dict:
    for _ in range(60):
        response = client.get(f"/api/v1/analytics/exports/{job_id}")
        body = response.json()
        if body["status"] in {"completed", "failed"}:
            return body
        time.sleep(0.05)
    raise AssertionError(f"export job did not finish in time: {job_id}")


def _make_settings(pipeline_names=("test_pipeline",)):
    pipelines = [
        PipelineConfig(
            name=name,
            table=name.upper(),
            source_where_template=(
                "DATE_COL >= :data_interval_start AND DATE_COL < :data_interval_end"
            ),
            strategy="full",
            schedule="0 2 * * *",
        )
        for name in pipeline_names
    ]
    settings = MagicMock()
    settings.pipelines = pipelines
    settings.api = ApiConfig(allowed_origins=["http://10.0.0.1"])
    settings.webui = WebUIConfig()
    return settings


def test_xlsx_writer_writes_workbook_with_header_and_rows(tmp_path):
    writer = XlsxExportWriter()
    artifact_path = tmp_path / "rows.xlsx"
    row_count = writer.write(artifact_path, _BatchReader([_build_test_batch(3)]))

    assert row_count == 3
    assert artifact_path.exists()

    workbook = load_workbook(artifact_path, read_only=True, data_only=True)
    try:
        worksheet = workbook["export"]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    assert rows[0] == ("source_name", "row_count")
    assert rows[1] == ("SRC_0", 0)
    assert rows[3] == ("SRC_2", 2)


def test_xlsx_writer_enforces_row_limit(tmp_path):
    writer = XlsxExportWriter(max_rows=2)
    artifact_path = tmp_path / "too-many.xlsx"

    with pytest.raises(XlsxRowLimitExceededError):
        writer.write(artifact_path, _BatchReader([_build_test_batch(3)]))

    assert not artifact_path.exists()


def test_export_service_creates_xlsx_artifact(tmp_path):
    export_service = FilesystemAnalyticsExportService(
        root_path=tmp_path / "exports",
        query_service=_StaticPlanQueryProvider([_build_test_batch(3)]),
    )
    try:
        create_response = export_service.create_export(
            ExportCreateRequest(
                dataset="mes_ops",
                action_key="xlsx_export",
                format=ExportFormat.XLSX,
                filters={},
            )
        )
        status = _wait_for_service_job(export_service, create_response.job_id)

        assert status.status is ExportJobStatus.COMPLETED
        assert status.file_name is not None
        assert status.file_name.endswith(".xlsx")
        assert status.row_count == 3
    finally:
        export_service.executor.shutdown(wait=True)


def test_export_service_marks_xlsx_job_failed_when_row_limit_exceeded(tmp_path):
    export_service = FilesystemAnalyticsExportService(
        root_path=tmp_path / "exports",
        query_service=_StaticPlanQueryProvider([_build_test_batch(3)]),
    )
    export_service._writers[ExportFormat.XLSX] = XlsxExportWriter(max_rows=2)
    try:
        create_response = export_service.create_export(
            ExportCreateRequest(
                dataset="mes_ops",
                action_key="xlsx_export",
                format=ExportFormat.XLSX,
                filters={},
            )
        )
        status = _wait_for_service_job(export_service, create_response.job_id)

        assert status.status is ExportJobStatus.FAILED
        assert status.error_message is not None
        assert "xlsx export row limit exceeded" in status.error_message
        assert "csv.zip or parquet" in status.error_message
    finally:
        export_service.executor.shutdown(wait=True)


def test_analytics_export_api_supports_xlsx(
    tmp_path,
    analytics_mart_builder,
):
    settings = _make_settings()
    run_repo = MagicMock()
    step_repo = MagicMock()
    run_repo.find_by_pipeline.return_value = []
    run_repo.find_by_pipeline_paginated.return_value = ([], 0)
    step_repo.find_by_pipeline_run.return_value = []

    database_path = tmp_path / "analytics.duckdb"
    analytics_mart_builder(database_path)
    query_service = DuckDBAnalyticsQueryService(database_path)
    export_service = FilesystemAnalyticsExportService(tmp_path / "exports", query_service)
    try:
        app = create_app(
            settings=settings,
            run_repo=run_repo,
            step_repo=step_repo,
            analytics_query_service=query_service,
            analytics_export_service=export_service,
        )
        client = TestClient(app)

        create_response = client.post(
            "/api/v1/analytics/exports",
            json={
                "dataset": "mes_ops",
                "action_key": "xlsx_export",
                "format": "xlsx",
                "filters": {"source": ["OPS_TABLE"]},
            },
        )

        assert create_response.status_code == 200
        job_id = create_response.json()["job_id"]
        status_body = _wait_for_api_job(client, job_id)
        assert status_body["status"] == "completed"
        assert status_body["file_name"].endswith(".xlsx")

        download_response = client.get(f"/api/v1/analytics/exports/{job_id}/download")
        assert download_response.status_code == 200
        assert ".xlsx" in download_response.headers.get("content-disposition", "")
    finally:
        export_service.executor.shutdown(wait=True)


def test_analytics_export_api_reports_failed_xlsx_when_row_limit_exceeded(
    tmp_path,
    analytics_mart_builder,
):
    settings = _make_settings()
    run_repo = MagicMock()
    step_repo = MagicMock()
    run_repo.find_by_pipeline.return_value = []
    run_repo.find_by_pipeline_paginated.return_value = ([], 0)
    step_repo.find_by_pipeline_run.return_value = []

    database_path = tmp_path / "analytics.duckdb"
    analytics_mart_builder(database_path)
    query_service = DuckDBAnalyticsQueryService(database_path)
    export_service = FilesystemAnalyticsExportService(tmp_path / "exports", query_service)
    export_service._writers[ExportFormat.XLSX] = XlsxExportWriter(max_rows=2)
    try:
        app = create_app(
            settings=settings,
            run_repo=run_repo,
            step_repo=step_repo,
            analytics_query_service=query_service,
            analytics_export_service=export_service,
        )
        client = TestClient(app)

        create_response = client.post(
            "/api/v1/analytics/exports",
            json={
                "dataset": "mes_ops",
                "action_key": "xlsx_export",
                "format": "xlsx",
                "filters": {},
            },
        )

        assert create_response.status_code == 200
        job_id = create_response.json()["job_id"]
        status_body = _wait_for_api_job(client, job_id)
        assert status_body["status"] == "failed"
        assert "xlsx export row limit exceeded" in status_body["error_message"]
    finally:
        export_service.executor.shutdown(wait=True)
